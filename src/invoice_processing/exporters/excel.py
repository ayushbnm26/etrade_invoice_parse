from __future__ import annotations

from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from invoice_processing.config import (
    EXCEPTION_COLUMNS,
    HEADER_COLUMNS,
    LINE_ITEM_COLUMNS,
    SUMMARY_COLUMNS,
    VALIDATION_COLUMNS,
)
from invoice_processing.exceptions import ExportError
from invoice_processing.models import BatchResult, ParsedInvoice


KNOWN_PUBLIC_TAX_TYPES = ("cgst", "sgst", "igst")
PUBLIC_ITEM_COLUMNS = [
    ("ASIN", "asin_code"),
    ("Quantity", "qty"),
    ("Price / Unit", "price_per_unit"),
    ("Net Amount", "net_amount"),
    ("CGST Rate", "cgst_rate"),
    ("CGST Amount", "cgst_amount"),
    ("SGST Rate", "sgst_rate"),
    ("SGST Amount", "sgst_amount"),
    ("IGST Rate", "igst_rate"),
    ("IGST Amount", "igst_amount"),
    ("Other Tax Type", "other_tax_type"),
    ("Other Tax Rate", "other_tax_rate"),
    ("Other Tax Amount", "other_tax_amount"),
    ("Total Amount", "total_amount"),
]
PUBLIC_RATE_FIELDS = {"cgst_rate", "sgst_rate", "igst_rate", "other_tax_rate"}
PUBLIC_AMOUNT_FIELDS = {
    "price_per_unit",
    "net_amount",
    "cgst_amount",
    "sgst_amount",
    "igst_amount",
    "other_tax_amount",
    "total_amount",
}


class ExcelExporter:
    def export(self, result: BatchResult, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        frames = {
            "Invoice_Header": self._frame(result.headers, HEADER_COLUMNS),
            "Invoice_Line_Items": self._frame(result.line_items, LINE_ITEM_COLUMNS),
            "Invoice_Summary": self._frame(result.summaries, SUMMARY_COLUMNS),
            "Validation": self._frame(result.validations, VALIDATION_COLUMNS),
            "Exceptions": self._frame(result.exceptions, EXCEPTION_COLUMNS),
        }

        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for sheet_name, frame in frames.items():
                    frame.to_excel(writer, sheet_name=sheet_name, index=False)

                for worksheet in writer.sheets.values():
                    worksheet.freeze_panes = "A2"
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
                            except Exception:
                                continue
                        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)
        except Exception as exc:
            raise ExportError(f"Unable to write Excel output '{output_path}': {exc}") from exc

        return output_path

    def export_public_invoice(self, parsed_invoice: ParsedInvoice, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            workbook = Workbook()
            details_sheet = workbook.active
            details_sheet.title = "Invoice Details"
            items_sheet = workbook.create_sheet("Items")

            self._build_public_details_sheet(details_sheet, parsed_invoice.header)
            self._build_public_items_sheet(
                items_sheet,
                parsed_invoice.header,
                parsed_invoice.line_items,
                parsed_invoice.summary,
            )

            workbook.save(output_path)
        except Exception as exc:
            raise ExportError(f"Unable to write public Excel output '{output_path}': {exc}") from exc

        return output_path

    def _frame(self, rows: list[dict[str, Any]], columns: list[str]) -> pd.DataFrame:
        frame = pd.DataFrame(rows)

        for column in columns:
            if column not in frame.columns:
                frame[column] = pd.NA

        extra_columns = [column for column in frame.columns if column not in columns]
        return frame[columns + extra_columns]

    def _build_public_details_sheet(self, worksheet: Any, header: dict[str, Any]) -> None:
        self._prepare_public_sheet(worksheet)

        accent_fill = PatternFill("solid", fgColor="EAF2F8")
        label_fill = PatternFill("solid", fgColor="F7F9FB")
        text_color = "1F2937"
        border = self._thin_border("D8E2E8")

        worksheet.merge_cells("A1:D1")
        title = worksheet["A1"]
        title.value = "Invoice Details"
        title.font = Font(size=16, bold=True, color=text_color)
        title.fill = accent_fill
        title.alignment = Alignment(horizontal="center", vertical="center")
        title.border = border
        worksheet.row_dimensions[1].height = 30

        detail_rows = [
            ("Invoice Number", header.get("invoice_number", "")),
            ("System Reference Number", header.get("system_ref_no", "")),
        ]

        row = 3
        for label, value in detail_rows:
            worksheet.cell(row=row, column=1, value=label)
            worksheet.cell(row=row, column=2, value=value)
            row += 1

        for cell_range in ("A3:B4",):
            for row_cells in worksheet[cell_range]:
                for cell in row_cells:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if cell.column == 1:
                        cell.fill = label_fill
                        cell.font = Font(bold=True, color=text_color)
                    else:
                        cell.font = Font(color=text_color)

        worksheet.merge_cells("A7:B7")
        worksheet.merge_cells("C7:D7")
        worksheet["A7"] = "Shipping Address"
        worksheet["C7"] = "Receiver Shipping Address"

        worksheet.merge_cells("A8:B14")
        worksheet.merge_cells("C8:D14")
        worksheet["A8"] = self._format_address(header, "shipping")
        worksheet["C8"] = self._format_address(header, "receiver_shipping")

        for cell_ref in ("A7", "C7"):
            cell = worksheet[cell_ref]
            cell.fill = accent_fill
            cell.font = Font(bold=True, color=text_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for cell_ref in ("A8", "C8"):
            cell = worksheet[cell_ref]
            cell.font = Font(color=text_color)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        for merged_range in ("A8:B14", "C8:D14"):
            for row_cells in worksheet[merged_range]:
                for cell in row_cells:
                    cell.border = border

        worksheet.column_dimensions["A"].width = 23
        worksheet.column_dimensions["B"].width = 28
        worksheet.column_dimensions["C"].width = 25
        worksheet.column_dimensions["D"].width = 28

    def _build_public_items_sheet(
        self,
        worksheet: Any,
        header: dict[str, Any],
        line_items: list[dict[str, Any]],
        summary: dict[str, Any],
    ) -> None:
        self._prepare_public_sheet(worksheet)

        accent_fill = PatternFill("solid", fgColor="EAF2F8")
        header_fill = PatternFill("solid", fgColor="DDEBF2")
        total_fill = PatternFill("solid", fgColor="F2F7F3")
        text_color = "1F2937"
        border = self._thin_border("D8E2E8")

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PUBLIC_ITEM_COLUMNS))
        title = worksheet.cell(row=1, column=1, value="Invoice Items")
        title.font = Font(size=16, bold=True, color=text_color)
        title.fill = accent_fill
        title.alignment = Alignment(horizontal="center", vertical="center")
        title.border = border
        worksheet.row_dimensions[1].height = 30

        header_row = 3
        for index, (label, _) in enumerate(PUBLIC_ITEM_COLUMNS, start=1):
            cell = worksheet.cell(row=header_row, column=index, value=label)
            cell.fill = header_fill
            cell.font = Font(bold=True, color=text_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        public_items = [self._public_item_record(item) for item in line_items]

        for row_index, item in enumerate(public_items, start=header_row + 1):
            for column_index, (_, field_name) in enumerate(PUBLIC_ITEM_COLUMNS, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=item.get(field_name, ""))
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(color=text_color)
                self._apply_public_number_format(cell, field_name)

        subtotal_row = header_row + len(line_items) + 1
        total_quantity_row = subtotal_row + 1
        subtotal = summary.get("subtotal_pdf") or self._sum_numeric(public_items, "net_amount")
        grand_total = summary.get("grand_total_pdf") or self._sum_numeric(public_items, "total_amount")
        total_quantity = summary.get("total_qty_pdf") or self._sum_numeric(public_items, "qty")

        worksheet.cell(row=subtotal_row, column=1, value="Sub Total")
        worksheet.cell(row=subtotal_row, column=4, value=subtotal)
        worksheet.cell(row=subtotal_row, column=6, value=self._sum_numeric_or_blank(public_items, "cgst_amount"))
        worksheet.cell(row=subtotal_row, column=8, value=self._sum_numeric_or_blank(public_items, "sgst_amount"))
        worksheet.cell(row=subtotal_row, column=10, value=self._sum_numeric_or_blank(public_items, "igst_amount"))
        worksheet.cell(row=subtotal_row, column=13, value=self._sum_numeric_or_blank(public_items, "other_tax_amount"))
        worksheet.cell(row=subtotal_row, column=14, value=grand_total)
        worksheet.cell(row=total_quantity_row, column=1, value="Total Quantity")
        worksheet.cell(row=total_quantity_row, column=2, value=total_quantity)

        for row_index in (subtotal_row, total_quantity_row):
            for column_index, (_, field_name) in enumerate(PUBLIC_ITEM_COLUMNS, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.fill = total_fill
                cell.border = border
                cell.font = Font(bold=column_index == 1, color=text_color)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                self._apply_public_number_format(cell, field_name)

        self._append_public_address_footer(
            worksheet,
            header,
            start_row=total_quantity_row + 2,
            border=border,
            text_color=text_color,
        )

        worksheet.freeze_panes = "A4"
        self._auto_fit_columns(worksheet, max_width=24)

    def _prepare_public_sheet(self, worksheet: Any) -> None:
        worksheet.sheet_view.showGridLines = False

    def _append_public_address_footer(
        self,
        worksheet: Any,
        header: dict[str, Any],
        *,
        start_row: int,
        border: Border,
        text_color: str,
    ) -> None:
        heading_fill = PatternFill("solid", fgColor="EAF2F8")
        body_fill = PatternFill("solid", fgColor="FBFCFE")
        body_start_row = start_row + 1
        body_end_row = start_row + 5
        address_blocks = [
            (1, 7, "Shipping Address", self._format_address(header, "shipping")),
            (
                8,
                len(PUBLIC_ITEM_COLUMNS),
                "Receiver Shipping Address",
                self._format_address(header, "receiver_shipping"),
            ),
        ]

        worksheet.row_dimensions[start_row].height = 22
        for row_index in range(body_start_row, body_end_row + 1):
            worksheet.row_dimensions[row_index].height = 20

        for start_column, end_column, label, address in address_blocks:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=start_column,
                end_row=start_row,
                end_column=end_column,
            )
            heading_cell = worksheet.cell(row=start_row, column=start_column, value=label)
            heading_cell.fill = heading_fill
            heading_cell.font = Font(bold=True, color=text_color)
            heading_cell.alignment = Alignment(horizontal="center", vertical="center")

            worksheet.merge_cells(
                start_row=body_start_row,
                start_column=start_column,
                end_row=body_end_row,
                end_column=end_column,
            )
            body_cell = worksheet.cell(
                row=body_start_row,
                column=start_column,
                value=address or "Not available",
            )
            body_cell.fill = body_fill
            body_cell.font = Font(color=text_color)
            body_cell.alignment = Alignment(vertical="top", wrap_text=True)

            for row_cells in worksheet.iter_rows(
                min_row=start_row,
                max_row=body_end_row,
                min_col=start_column,
                max_col=end_column,
            ):
                for cell in row_cells:
                    cell.border = border
                    if cell.row > start_row:
                        cell.fill = body_fill

    def _format_address(self, header: dict[str, Any], prefix: str) -> str:
        lines = [
            header.get(f"{prefix}_name", ""),
            header.get(f"{prefix}_address", ""),
        ]

        state_code = header.get(f"{prefix}_state_code", "")
        gstin = header.get(f"{prefix}_gstin", "")
        pan = header.get(f"{prefix}_pan", "")

        if state_code:
            lines.append(f"State Code: {state_code}")
        if gstin:
            lines.append(f"GSTIN: {gstin}")
        if pan:
            lines.append(f"PAN: {pan}")

        return "\n".join(str(line) for line in lines if line)

    def _public_item_record(self, item: dict[str, Any]) -> dict[str, Any]:
        public_item = dict(item)
        tax_type_raw = str(item.get("tax_type_raw") or "").upper()

        for tax_type in KNOWN_PUBLIC_TAX_TYPES:
            if public_item.get(f"{tax_type}_rate") is None and tax_type.upper() in tax_type_raw:
                public_item[f"{tax_type}_rate"] = item.get("tax_rate_raw")
            if public_item.get(f"{tax_type}_amount") is None and tax_type.upper() in tax_type_raw:
                public_item[f"{tax_type}_amount"] = item.get("tax_amount_raw")

        tax_tokens = re.findall(r"[A-Z][A-Z0-9]*", tax_type_raw)
        other_tax_tokens = [
            token for token in tax_tokens if token not in {tax_type.upper() for tax_type in KNOWN_PUBLIC_TAX_TYPES}
        ]
        has_known_tax_amount = any(public_item.get(f"{tax_type}_amount") is not None for tax_type in KNOWN_PUBLIC_TAX_TYPES)

        public_item["other_tax_type"] = " ".join(other_tax_tokens)
        public_item["other_tax_rate"] = item.get("tax_rate_raw") if other_tax_tokens and not has_known_tax_amount else None
        public_item["other_tax_amount"] = item.get("tax_amount_raw") if other_tax_tokens and not has_known_tax_amount else None

        return public_item

    def _apply_public_number_format(self, cell: Any, field_name: str) -> None:
        if field_name == "qty":
            cell.number_format = "#,##0"
        elif field_name in PUBLIC_AMOUNT_FIELDS:
            cell.number_format = "#,##0.00"
        elif field_name in PUBLIC_RATE_FIELDS:
            cell.number_format = "0.00"

    def _auto_fit_columns(self, worksheet: Any, max_width: int = 60) -> None:
        for column_index, column in enumerate(worksheet.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(column_index)
            for cell in column:
                value = cell.value
                if value is None:
                    continue
                parts = str(value).splitlines() or [""]
                max_length = max(max_length, *(len(part) for part in parts))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), max_width)

    def _thin_border(self, color: str) -> Border:
        side = Side(style="thin", color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    def _sum_numeric(self, rows: list[dict[str, Any]], key: str) -> float:
        total = 0.0
        for row in rows:
            value = row.get(key)
            if isinstance(value, int | float):
                total += float(value)
        return total

    def _sum_numeric_or_blank(self, rows: list[dict[str, Any]], key: str) -> float | None:
        values = [float(row[key]) for row in rows if isinstance(row.get(key), int | float)]
        return sum(values) if values else None
