from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from invoice_processing.exporters.excel import ExcelExporter  # noqa: E402
from invoice_processing.models import ParsedInvoice  # noqa: E402


class PublicWorkbookExporterTests(unittest.TestCase):
    def test_public_workbook_has_requested_sheets_and_totals(self) -> None:
        parsed_invoice = ParsedInvoice(
            header={
                "invoice_number": "INV-1001",
                "system_ref_no": "SYS-1001",
                "shipping_name": "Seller Warehouse",
                "shipping_address": "Mumbai, Maharashtra",
                "receiver_shipping_name": "Customer Receiving",
                "receiver_shipping_address": "Bengaluru, Karnataka",
            },
            summary={
                "subtotal_pdf": 250.0,
                "grand_total_pdf": 295.0,
                "total_qty_pdf": 3,
            },
            line_items=[
                {
                    "asin_code": "B000TEST",
                    "qty": 3,
                    "price_per_unit": 83.33,
                    "net_amount": 250.0,
                    "tax_rate_raw": "9\n9",
                    "tax_type_raw": "CGST SGST",
                    "tax_amount_raw": "22.5\n22.5",
                    "cgst_rate": 9.0,
                    "cgst_amount": 22.5,
                    "sgst_rate": 9.0,
                    "sgst_amount": 22.5,
                    "igst_rate": None,
                    "igst_amount": None,
                    "total_amount": 295.0,
                }
            ],
            validation={},
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "invoice_public.xlsx"
            ExcelExporter().export_public_invoice(parsed_invoice, output_path)

            workbook = load_workbook(output_path)
            self.assertEqual(workbook.sheetnames, ["Invoice Details", "Items"])

            details = workbook["Invoice Details"]
            self.assertEqual(details["A3"].value, "Invoice Number")
            self.assertEqual(details["B3"].value, "INV-1001")
            self.assertEqual(details["A4"].value, "System Reference Number")
            self.assertEqual(details["B4"].value, "SYS-1001")
            self.assertIn("Seller Warehouse", details["A8"].value)
            self.assertIn("Customer Receiving", details["C8"].value)

            items = workbook["Items"]
            headers = [items.cell(row=3, column=column).value for column in range(1, 15)]
            self.assertEqual(
                headers,
                [
                    "ASIN",
                    "Quantity",
                    "Price / Unit",
                    "Net Amount",
                    "CGST Rate",
                    "CGST Amount",
                    "SGST Rate",
                    "SGST Amount",
                    "IGST Rate",
                    "IGST Amount",
                    "Other Tax Type",
                    "Other Tax Rate",
                    "Other Tax Amount",
                    "Total Amount",
                ],
            )
            self.assertEqual(items["E4"].value, 9.0)
            self.assertEqual(items["F4"].value, 22.5)
            self.assertEqual(items["G4"].value, 9.0)
            self.assertEqual(items["H4"].value, 22.5)
            self.assertIsNone(items["I4"].value)
            self.assertIsNone(items["J4"].value)
            self.assertEqual(items["A5"].value, "Sub Total")
            self.assertEqual(items["D5"].value, 250.0)
            self.assertEqual(items["F5"].value, 22.5)
            self.assertEqual(items["H5"].value, 22.5)
            self.assertIsNone(items["J5"].value)
            self.assertEqual(items["N5"].value, 295.0)
            self.assertEqual(items["A6"].value, "Total Quantity")
            self.assertEqual(items["B6"].value, 3)
            self.assertEqual(items["A8"].value, "Shipping Address")
            self.assertEqual(items["H8"].value, "Receiver Shipping Address")
            self.assertIn("Seller Warehouse", items["A9"].value)
            self.assertIn("Mumbai, Maharashtra", items["A9"].value)
            self.assertIn("Customer Receiving", items["H9"].value)
            self.assertIn("Bengaluru, Karnataka", items["H9"].value)


if __name__ == "__main__":
    unittest.main()
